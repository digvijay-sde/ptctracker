from flask import Flask, render_template, request, redirect, session
import openpyxl

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# Function to read the users' login data from the Excel file
def read_users():
    wb = openpyxl.load_workbook('delivery_data.xlsx')
    sheet = wb['users']  # Accessing the 'users' sheet

    users = {}
    for row in range(2, sheet.max_row + 1):  # Skipping the header row
        user_id = sheet.cell(row=row, column=1).value
        username = sheet.cell(row=row, column=2).value
        password = sheet.cell(row=row, column=3).value
        users[username] = {'user_id': user_id, 'password': password}
    
    return users

# Function to read delivery data from the Excel file
def read_deliveries():
    wb = openpyxl.load_workbook('delivery_data.xlsx')
    sheet = wb['deliveries']  # Accessing the 'deliveries' sheet

    deliveries = []
    for row in range(2, sheet.max_row + 1):  # Skipping the header row
        delivery = {
            'User ID': sheet.cell(row=row, column=1).value,
            'Delivery Date': sheet.cell(row=row, column=2).value,
            'Delivery Image URL': sheet.cell(row=row, column=3).value,
            'Delivery Number': sheet.cell(row=row, column=4).value,
            'Weight': sheet.cell(row=row, column=5).value,
            'Items': sheet.cell(row=row, column=6).value,
            'Transport Name': sheet.cell(row=row, column=7).value,
        }
        deliveries.append(delivery)
    return deliveries

# Login route
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        users = read_users()
        
        if username in users and users[username]['password'] == password:
            session['username'] = username
            session['user_id'] = users[username]['user_id']
            return redirect('/dashboard')
        else:
            return "Invalid credentials. Try again."
    
    return render_template('login.html')

# Dashboard route
@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect('/')

    user_id = session['user_id']
    deliveries = read_deliveries()

    # Filter deliveries for the logged-in user
    user_deliveries = [d for d in deliveries if d['User ID'] == user_id]

    return render_template('dashboard.html', deliveries=user_deliveries)

# Logout route
@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

if __name__ == '__main__':
    app.run(debug=True)
